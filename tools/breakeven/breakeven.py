from pyomo.environ import *
from pyomo.core import Constraint
from pyomo.opt import SolverFactory
import sys, os, platform
from matplotlib import pyplot as plt
import numpy as np
from collections import OrderedDict, defaultdict
from time import time
import pandas as pd
from IPython import embed as IP
from temoa_model import temoa_create_model
from pyomo.core.base import Var, Constraint, Objective, maximize, minimize
from pyomo.repn import generate_canonical_repn

import cplex, CplexSolverError
from openpyxl import Workbook
import cplex

def return_Temoa_model():

    model = temoa_create_model()

    model.dual  = Suffix(direction=Suffix.IMPORT)
    model.rc    = Suffix(direction=Suffix.IMPORT)
    model.slack = Suffix(direction=Suffix.IMPORT)
    model.lrc   = Suffix(direction=Suffix.IMPORT)
    model.urc   = Suffix(direction=Suffix.IMPORT)
    return model

def return_Temoa_data(model, list_dat):
    data = DataPortal(model = model)
    for d in list_dat:
        data.load(filename=d)
    return data

def return_c_vector(block, unfixed):
    # Note that this function is adapted function collect_linear_terms defined
    # in pyomo/repn/collect.py.
    from pyutilib.misc import Bunch

    #
    # Variables are constraints of block
    # Constraints are unfixed variables of block and the parent model.
    #
    vnames = set()
    for (name, data) in block.component_map(Constraint, active=True).items():
        vnames.add((name, data.is_indexed()))
    cnames = set(unfixed)
    for (name, data) in block.component_map(Var, active=True).items():
        cnames.add((name, data.is_indexed()))
    #
    A = {}
    b_coef = {}
    c_rhs = {}
    c_sense = {}
    d_sense = None
    v_domain = {}
    #
    # Collect objective
    #
    for (oname, odata) in block.component_map(Objective, active=True).items():
        for ndx in odata:
            if odata[ndx].sense == maximize:
                o_terms = generate_canonical_repn(-1*odata[ndx].expr, compute_values=False)
                d_sense = minimize
            else:
                o_terms = generate_canonical_repn(odata[ndx].expr, compute_values=False)
                d_sense = maximize
            for i in range(len(o_terms.variables)):
                c_rhs[ o_terms.variables[i].parent_component().local_name, o_terms.variables[i].index() ] = o_terms.linear[i]
        # Stop after the first objective
        break
    return c_rhs

def coef_IC(instance, target_tech, target_year):
    # This function returns coefficient associated with IC given t, v
    t = target_tech
    v = target_year
    P_0 = min( instance.time_optimize )
    P_e = instance.time_future.last()
    GDR = value( instance.GlobalDiscountRate )
    MPL = instance.ModelProcessLife
    LLN = instance.LifetimeLoanProcess
    x   = 1 + GDR    # convenience variable, nothing more.
    period_available = set()
    for p in instance.time_future:
        if (p, t, v) in instance.CostFixed.keys():
            period_available.add(p)
    c_i = ( 
            instance.CostInvest[t, v] 
            * instance.LoanAnnualize[t, v] 
            * ( LLN[t, v] if not GDR else 
                (x**(P_0 - v + 1) 
                * ( 1 - x **( -value(LLN[t, v]) ) ) 
                / GDR) 
                ) 
    ) * (
		  (
			  1 -  x**( -min( value(instance.LifetimeProcess[t, v]), P_e - v ) )
		  )
		  /(
			  1 -  x**( -value( instance.LifetimeProcess[t, v] ) ) 
		  )
    )
    return value(c_i)

def coef_FC(instance, target_tech, target_year):
    # This function returns coefficient associated with FC given t, v
    t = target_tech
    v = target_year
    P_0 = min( instance.time_optimize )
    P_e = instance.time_future.last()
    GDR = value( instance.GlobalDiscountRate )
    MPL = instance.ModelProcessLife
    LLN = instance.LifetimeLoanProcess
    x   = 1 + GDR    # convenience variable, nothing more.
    period_available = set()
    for p in instance.time_future:
        if (p, t, v) in instance.CostFixed.keys():
            period_available.add(p)

    c_f = sum( 
        instance.CostFixed[p, t, v]
        * ( MPL[p, t, v] if not GDR else
            (x**(P_0 - p + 1)
            * ( 1 - x**( -value(MPL[p, t, v]) ) )
            / GDR ) 
            )
        for p in period_available 
    ) 
    return value(c_f)

def validate_coef(c0, instance, target_tech, target_year):
    # This function validates if c0 equals the correct coefficient of process
    # (target_tech, target_year)
    c_i = coef_IC(instance, target_tech, target_year)
    c_f = coef_FC(instance, target_tech, target_year)
    c = c_i + c_f
    if value(c - c0) <= 1E-5: # Compatible with Pyomo 5.5
        return True
    else:
        return False

def sensitivity(dat, techs):
    # This function performs break-even analysis for technologies specified in 
    # the argument techs. It uses suffix from Pyomo and returns the breakeven 
    # cost as screen outputs. Note that the Pyomo suffix sometimes returns 
    # anomalous values, and that's why I create another function, 
    # sensitivity_api() to use Python API for CPLEX.

    model = temoa_create_model()
    
    model.dual  = Suffix(direction=Suffix.IMPORT)
    model.rc    = Suffix(direction=Suffix.IMPORT)
    model.slack = Suffix(direction=Suffix.IMPORT)
    model.lrc   = Suffix(direction=Suffix.IMPORT)
    model.urc   = Suffix(direction=Suffix.IMPORT)

    data = DataPortal(model = model)
    for d in dat:
        data.load(filename=d)
    instance = model.create_instance(data)
    optimizer = SolverFactory('cplex')
    optimizer.options['lpmethod'] = 1 # Use primal simplex
    results = optimizer.solve(instance, suffixes=['dual', 'urc', 'slack', 'lrc'])
    instance.solutions.load_from(results)

    coef_CAP = dict()
    scal_CAP = dict()
    # Break-even investment cost for this scenario, indexed by technology
    years    = list()
    bic_s    = dict()
    ic_s     = dict() # Raw investment costs for this scenario, indexed by tech
    cap_s    = dict()
    for t in techs:
        vintages = instance.vintage_optimize
        P_0 = min( instance.time_optimize )
        GDR = value( instance.GlobalDiscountRate )
        MPL = instance.ModelProcessLife
        LLN = instance.LifetimeLoanProcess
        x   = 1 + GDR    # convenience variable, nothing more.

        bic_s[t] = list()
        ic_s[t]  = list()
        cap_s[t] = list()
        years = vintages.value
        for v in vintages:
            period_available = set()
            for p in instance.time_future:
                if (p, t, v) in instance.CostFixed.keys():
                    period_available.add(p)
            c_i = ( 
                    instance.CostInvest[t, v] 
                    * instance.LoanAnnualize[t, v] 
                    * ( LLN[t, v] if not GDR else 
                        (x**(P_0 - v + 1) 
                        * ( 1 - x **( -value(LLN[t, v]) ) ) 
                        / GDR) 
                      ) 
            )

            c_s = (-1)*(
                value( instance.CostInvest[t, v] )
                * value( instance.SalvageRate[t, v] )
                / ( 1 if not GDR else 
                    (1 + GDR)**( 
                        instance.time_future.last() 
                        - instance.time_future.first()
                        - 1
                        ) 
                    )
                )

            c_f = sum( 
                instance.CostFixed[p, t, v]
                * ( MPL[p, t, v] if not GDR else
                    (x**(P_0 - p + 1)
                    * ( 1 - x**( -value(MPL[p, t, v]) ) )
                    / GDR ) 
                  )
                for p in period_available 
            ) 

            c = c_i + c_s + c_f
            s = (c - instance.lrc[instance.V_Capacity[t, v]])/c
            coef_CAP[t, v] = c
            scal_CAP[t, v] = s # Must reduce TO this percentage
            bic_s[t].append(scal_CAP[t, v]*instance.CostInvest[t, v])
            ic_s[t].append(instance.CostInvest[t, v])
            cap_s[t].append( value( instance.V_Capacity[t, v] ) )

        # print("Tech\tVintage\tL. RC\tCoef\tU .RC\tScale\tBE IC\tBE FC\tIC\tFC\tCap")
        print("{:>10s}\t{:>7s}\t{:>6s}\t{:>4s}\t{:>6s}\t{:>5s}\t{:>7s}\t{:>7s}\t{:>5s}\t{:>3s}\t{:>5s}".format('Tech','Vintage', 'L. RC', 'Coef', 'U. RC', 'Scale', 'BE IC', 'BE FC', 'IC', 'FC', 'Cap'))
        for v in vintages:
            lrc = instance.lrc[instance.V_Capacity[t, v]]
            urc = instance.urc[instance.V_Capacity[t, v]]

            # print("{:>s}\t{:>g}\t{:>.0f}\t{:>.0f}\t{:>.0f}\t{:>.3f}\t{:>.1f}\t{:>.1f}\t{:>.0f}\t{:>.0f}\t{:>.3f}".format()
            print("{:>10s}\t{:>7g}\t{:>6.0f}\t{:>4.0f}\t{:>6.0f}\t{:>5.3f}\t{:>7.1f}\t{:>7.1f}\t{:>5.0f}\t{:>3.0f}\t{:>5.3f}".format(
            t, v, lrc, coef_CAP[t, v], urc, scal_CAP[t, v], 
            scal_CAP[t, v]*instance.CostInvest[t, v], 
            scal_CAP[t, v]*instance.CostFixed[v, t, v], # Use the FC of the first period
            instance.CostInvest[t,v],
            instance.CostFixed[v, t, v],
            value(instance.V_Capacity[t, v])
            ))

    print('Dual and slack variables for emission caps:')
    for e in instance.commodity_emissions:
        for p in instance.time_optimize:
            if (p, e) in instance.EmissionLimitConstraint:
                print(p, e, instance.dual[instance.EmissionLimitConstraint[p, e]], '\t', instance.slack[instance.EmissionLimitConstraint[p, e]])
    return years, bic_s, ic_s

    print('Dual and slack variables for Commodity Demand Constraints')
    for c in instance.commodity_demand:
        for p in instance.time_optimize:
            for s in instance.time_season:
                for tod in instance.time_of_day:
                    print(p, s, tod, instance.dual[instance.DemandConstraint[p,s,tod,c]], instance.slack[instance.DemandConstraint[p,s,tod,c]])

def sensitivity_api(instance, techs, algorithm=None):
    # This code block realizes the same function as sensitivity(), however 
    # because I am using Python API for CPLEX here, it only works when the 
    # solver is CPLEX. I also updated the returned value and now it is a pandas 
    # DataFramework, which supports fast csv creation.

    instance.write('tmp.lp', io_options={'symbolic_solver_labels':True})
    c = cplex.Cplex('tmp.lp')
    os.remove('tmp.lp')
    c.set_results_stream(None) # Turn screen output off

    msg = ''
    if algorithm:
        if algorithm == "o":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.auto)
        elif algorithm == "p":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.primal)
        elif algorithm == "d":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.dual)
        elif algorithm == "b":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.barrier)
            c.parameters.barrier.crossover.set(
                c.parameters.barrier.crossover.values.none)
        elif algorithm == "h":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.barrier)
        elif algorithm == "s":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.sifting)
        elif algorithm == "c":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.concurrent)
        else:
            raise ValueError(
                'method must be one of "o", "p", "d", "b", "h", "s" or "c"')

    try:
        c.solve()
    except CplexSolverError:
        print("Exception raised during solve")
        return

    vintages = list(instance.vintage_optimize)
    coef_CAP = dict()
    scal_CAP = dict()
    # Break-even investment cost for this scenario, indexed by technology
    years    = list()
    bic_s    = dict()
    ic_s     = dict() # Raw investment costs for this scenario, indexed by tech
    cap_s    = dict()
    clb_s    = dict()
    cub_s    = dict()
    results  = list()
    for t in techs:
        bic_s[t] = list()
        ic_s[t]  = list()
        cap_s[t] = list()
        for v in vintages:
            target_var  = 'V_Capacity(' + t + '_' + str(v) + ')'
            c0          = c.objective.get_linear(target_var)
            clb, cub    = c.solution.sensitivity.objective(target_var) # Coefficient lower bound, coefficient upper bound
            if cub > 1E5:
                cub = 0 # Infinity
            clb_s[t, v], cub_s[t, v] = clb, cub
            c_i = coef_IC(instance, t, v)
            c_f = coef_FC(instance, t, v)
            if not validate_coef(c0, instance, t, v):
                print('Error: Check coefficients!')
                sys.exit(0)
            coef_CAP[t, v] = c0
            scal_CAP[t, v] = clb/c0 # Break-even cost 1: Scaling both IC and FC
            alpha = c_i/value(instance.CostInvest[t, v])
            bic   = (clb - c_f)/alpha # Break-even cost 2: Only decrease IC
            bic_s[t].append(bic)
            ic_s[t].append(value(instance.CostInvest[t, v]))
            cap_s[t].append( c.solution.get_values(target_var) )

        print("{:>10s}\t{:>7s}\t{:>6s}\t{:>4s}\t{:>6s}\t{:>5s}\t{:>7s}\t{:>5s}\t{:>5s}".format(
            'Tech','Vintage', 'L. CB', 'Coef', 'U. CB', 'Scale', 'BE IC', 'IC', 'Cap',
        ))
        msg += "{:>10s}\t{:>7s}\t{:>6s}\t{:>4s}\t{:>6s}\t{:>5s}\t{:>7s}\t{:>5s}\t{:>5s}".format(
            'Tech','Vintage', 'L. CB', 'Coef', 'U. CB', 'Scale', 'BE IC', 'IC', 'Cap',
        )
        msg == '\n'
        for v in vintages:
            deployed = abs(cap_s[t][vintages.index(v)]) >= 1E-3
            tmp_beic_cs = value(instance.CostInvest[t, v]) if deployed else bic_s[t][vintages.index(v)]
            tmp_bes_cs  = 1 if deployed else scal_CAP[t, v]
            row = {
                'algorithm':         algorithm,
                'scenario':          None,
                'technology':        t,
                'vintage':           v,
                'coef lower bound':  clb_s[t, v],
                'coefficient':       coef_CAP[t, v],
                'coef upper bound':  cub_s[t, v],
                'scale':             scal_CAP[t, v],
                'BE IC':             bic_s[t][vintages.index(v)],
                'IC':                value(instance.CostInvest[t,v]),
                'capacity':          cap_s[t][vintages.index(v)],
                'BE IC (CS)':        tmp_beic_cs, 
                'scale (CS)':        tmp_bes_cs,
            }
            results.append(row)
            print("{:>10s}\t{:>7g}\t{:>6.0f}\t{:>4.0f}\t{:>6.0f}\t{:>5.3f}\t{:>7.1f}\t{:>5.0f}\t{:>5.3f}".format(
            t,
            v, 
            clb_s[t, v],
            coef_CAP[t, v],
            cub_s[t, v],
            scal_CAP[t, v],
            bic_s[t][vintages.index(v)],
            value(instance.CostInvest[t,v]),
            cap_s[t][vintages.index(v)]
            ))

            msg += "{:>10s}\t{:>7g}\t{:>6.0f}\t{:>4.0f}\t{:>6.0f}\t{:>5.3f}\t{:>7.1f}\t{:>5.0f}\t{:>5.3f}".format(
            t,
            v, 
            clb_s[t, v],
            coef_CAP[t, v],
            cub_s[t, v],
            scal_CAP[t, v],
            bic_s[t][vintages.index(v)],
            value(instance.CostInvest[t,v]),
            cap_s[t][vintages.index(v)]
            )
            msg += '\n'
    
    return msg, pd.DataFrame(results)

def bin_search(tech, vintage, dat, eps = 0.01, all_v = False):
    # This code block performs breakeven analysis by brutal force in an 
    # iterative way. I did this because sometimes sensitivity() and 
    # sensitivity_api() return anomalous values. Note that this code block 
    # returns the absolutely correct breakeven costs, however, it is 
    # significantly more time-consuming, since it takes 8-9 instances to 
    # calculate the breakeven cost of just one technology of one vintage.
    # Arguments are defined below:
    # tech     -> Target technology.
    # vintage  -> Target vintage. It is break-even when capacity in this year >= 0 
    # dat      -> A list of .dat files.
    # eps      -> Convergence tolerance
    # all_v    -> A flag used indicate the costs of which vintages are subject 
    # to change. If it is FALSE, then only the investment costs and fixed costs
    # in the target vintage will be altered, otherwise all vintages are affected
    # Note that, only the capacity of the target vintage will be monitored and 
    # be used to signal breakeven.
    monitor_year = vintage
    monitor_tech = tech

    t0 = time()
    time_mark = lambda: time() - t0 

    model = return_Temoa_model()
    optimizer = SolverFactory('cplex')
    data = return_Temoa_data(model, dat)
    instance = model.create_instance(data)

    time_optimize = [ i for i in data['time_future'] ]
    time_optimize.sort()
    ic0 = dict()
    fc0 = dict()
    if all_v:
        for v in time_optimize:
            if (monitor_tech, v) in data['CostInvest']:
                ic0[monitor_tech, v] = data['CostInvest'][monitor_tech, v]
                for p in time_optimize:
                    if (p, monitor_tech, v) in data['CostFixed']:
                        fc0[p, monitor_tech, v] = data['CostFixed'][p, monitor_tech, v]
    else:
        ic0[monitor_tech, monitor_year] = data['CostInvest'][monitor_tech, monitor_year]
        for p in time_optimize:
            if (p, monitor_tech, monitor_year) in data['CostFixed']:
                fc0[p, monitor_tech, monitor_year] = data['CostFixed'][p, monitor_tech, monitor_year]

    cap_target = 0
    scale_u = 1.0
    scale_l = 0.0

    history = dict()
    history['scale_u'] = [scale_u]
    history['scale_l'] = [scale_l]

    counter = 0
    scale_this = scale_u # Starting scale

    print('Iteration # {} starts at {} s'.format( counter, time_mark() ))
    instance = model.create_instance(data)
    instance.preprocess()
    results = optimizer.solve(instance, suffixes=['dual', 'urc', 'slack', 'lrc'])
    instance.solutions.load_from(results)
    cap_target = value( instance.V_Capacity[monitor_tech, monitor_year] )
    print('Iteration # {} solved at {} s'.format( counter, time_mark() ))
    print('Iteration # {}, scale: {:1.2f}, capacity: {} GW'.format( 
        counter,
        scale_this,
        cap_target
    ))
    if 1.0 - scale_this <= eps and cap_target > 0:
        return scale_this

    while (scale_u - scale_l) >= eps and counter <= 20:
        if cap_target <= 0:
            scale_u = scale_this
            history['scale_u'].append(scale_u)
        else:
            scale_l = scale_this
            history['scale_l'].append(scale_l)
        counter += 1

        scale_this = (scale_u + scale_l)*0.5
        for k in ic0:
            data['CostInvest'][k] = scale_this*ic0[k]
        for k in fc0:
            data['CostFixed'][k] = scale_this*fc0[k]

        print('Iteration # {} starts at {} s'.format( counter, time_mark() ))
        instance = model.create_instance(data)
        instance.preprocess()
        results = optimizer.solve(instance, suffixes=['dual', 'urc', 'slack', 'lrc'])
        instance.solutions.load_from(results)
        cap_target = value( instance.V_Capacity[monitor_tech, monitor_year] )
        print('Iteration # {} solved at {} s'.format( counter, time_mark() ))
        print('Iteration # {}, scale: {:1.2f}, capacity: {} GW'.format( 
            counter,
            scale_this,
            cap_target))
    return (scale_u + scale_l)/2.0

def cplex_search(t, v, cplex_instance, eps = 0.01, all_v=False):
    # This code block performs breakeven analysis by brutal force in an 
    # iterative way. However, it differs from bin_search() in that this function 
    # saves instance creation time by modifying the LP model coefficients in the 
    # CPLEX instance directly. By contrast, bin_search() has to create a new lp 
    # file each iteration after model coefficients are updated, which is 
    # time-consuming since Pyomo is notorious slow in model instantiation and 
    # creation. Other than that, these two functions are the same.
    # Arguments are defined below:
    # tech     -> Target technology.
    # vintage  -> Target vintage. It is break-even when capacity in this year >= 0 
    # dat      -> A list of .dat files.
    # eps      -> Convergence tolerance
    # all_v    -> A flag used indicate the costs of which vintages are subject 
    # to change. If it is FALSE, then only the investment costs and fixed costs
    # in the target vintage will be altered, otherwise all vintages are affected
    # Note that, only the capacity of the target vintage will be monitored and 
    # be used to signal breakeven.

    def return_row(c0, scale, capacity):
        row = {
            'algorithm':         cplex_instance.parameters.lpmethod,
            'scenario':          None,
            'technology':        t,
            'vintage':           v,
            'coef lower bound':  'N/A',
            'coefficient':       c0,
            'coef upper bound':  'N/A',
            'scale':             scale,
            'BE IC':             None, 
            'BE FC':             None,
            'IC':                None,
            'FC':                None,
            'capacity':          capacity,
            'BE IC (CS)':        'N/A', 
            'BE FC (CS)':        'N/A',
            'scale (CS)':        'N/A',
        }
        return row

    msg = ''
    target_year = v
    target_tech = t
    target_var0 = 'V_Capacity(' + target_tech + '_' + str(target_year) + ')'

    t0 = time()
    time_mark = lambda: time() - t0

    c0 = cplex_instance.objective.get_linear(target_var0) # Original coefficient
    cplex_instance.set_results_stream(None)

    scale_u = 1.0
    scale_l = 0.0

    history = dict()
    history['scale_u'] = [scale_u]
    history['scale_l'] = [scale_l]

    counter = 0
    scale_this = scale_u # Starting scale

    print('Iteration # {} starts at {} s'.format( counter, time_mark() ))
    msg += 'Iteration # {} starts at {} s\n'.format( counter, time_mark() )
    try:
        cplex_instance.solve()
    except CplexSolverError:
        print("Exception raised during solve")
        msg += "Exception raised during solve\n"
        return msg, None
    cap_target0 = cplex_instance.solution.get_values(target_var0)
    print('Iteration # {} solved at {} s'.format( counter, time_mark() ))
    msg += 'Iteration # {} solved at {} s\n'.format( counter, time_mark() )
    print('Iteration # {}, scale: {:1.2f}, capacity: {} GW'.format( 
        counter,
        scale_this,
        cap_target0
    ))
    if 1.0 - scale_this <= eps and cap_target0 > 0:
        row = return_row(c0, scale_this, cap_target0)
        return msg, pd.DataFrame([row])

    cap_target = cap_target0
    while (scale_u - scale_l) >= eps and counter <= 20:
        if cap_target <= 0:
            scale_u = scale_this
            history['scale_u'].append(scale_u)
        else:
            scale_l = scale_this
            history['scale_l'].append(scale_l)
        counter += 1

        scale_this = (scale_u + scale_l)*0.5
        cplex_instance.objective.set_linear(target_var0, scale_this*c0)

        print('Iteration # {} starts at {} s'.format( counter, time_mark() ))
        msg += 'Iteration # {} starts at {} s'.format( counter, time_mark() )
        try:
            cplex_instance.solve()
        except CplexSolverError:
            print("Exception raised during solve")
            msg += "Exception raised during solve\n"
            return msg, None
        cap_target = cplex_instance.solution.get_values(target_var0)
        print('Iteration # {} solved at {} s'.format( counter, time_mark() ))
        msg += 'Iteration # {} solved at {} s\n'.format( counter, time_mark() )
        print('Iteration # {}, scale: {:1.2f}, capacity: {} GW'.format( 
            counter,
            scale_this,
            cap_target))
    row = return_row(c0, scale_this, cap_target0)
    return msg, pd.DataFrame([row])

def sen_range_api(tech, vintage, scales, list_dat):
    # This function is adapted from CPLEX's example script lpex2.py
    # It does the same thing as sen_range, but with CPLEX API for Python

    # Given a range of scaling factor for coefficient of a specific V_Capacity, 
    # returns objective value, reduced cost, capacity etc. for each scaling 
    # factor

    target_year = vintage
    target_tech = tech
    target_var0 = 'V_Capacity(' + target_tech + '_' + str(target_year) + ')'
    algmap = {
        'primal simplex': 'p',
        'dual simplex':   'd',
        'barrier':        'h', # This is cross-over mode, since pure interior causes problems
        'default':        'o',
    } # cplex definition

    t0 = time()
    time_mark = lambda: time() - t0

    model = return_Temoa_model()
    data = return_Temoa_data(model, list_dat)
    instance = model.create_instance(data)

    ic0         = data['CostInvest'][target_tech, target_year]
    fc0         = data['CostFixed'][target_year, target_tech, target_year]
    all_periods = data['time_future']

    obj  = dict()
    cap  = dict()
    coef = dict()
    bic  = dict()
    bfc  = dict()
    ic   = dict() # Original IC
    fc   = dict() # Original FC
    clb  = dict() # Lower bound of objective coefficient
    cub  = dict() # Upper bound of objective coefficient
    rc   = dict() # Reduced cost

    for algorithm in ['barrier', 'dual simplex', 'primal simplex']:
        print('Algorithm: {}'.format( algorithm ))
        instance.write('tmp.lp', io_options={'symbolic_solver_labels':True})
        c = cplex.Cplex('tmp.lp')
        os.remove('tmp.lp')
        c.set_results_stream(None) # Turn screen output off
        c0 = c.objective.get_linear(target_var0)
        if not validate_coef(c0, instance, target_tech, target_year):
            print('Error!')
            sys.exit(0)
        print('[{:>9.2f}] CPLEX model loaded.'.format( time_mark() ))

        if algmap[algorithm] == "o":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.auto)
        elif algmap[algorithm] == "p":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.primal)
        elif algmap[algorithm] == "d":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.dual)
        elif algmap[algorithm] == "b":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.barrier)
            c.parameters.barrier.crossover.set(
                c.parameters.barrier.crossover.values.none)
        elif algmap[algorithm] == "h":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.barrier)
        elif algmap[algorithm] == "s":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.sifting)
        elif algmap[algorithm] == "c":
            c.parameters.lpmethod.set(c.parameters.lpmethod.values.concurrent)
        else:
            raise ValueError(
                'method must be one of "o", "p", "d", "b", "h", "s" or "c"')

        obj_alg  = list()
        cap_alg  = defaultdict(list)
        coef_alg = defaultdict(list)
        bic_alg  = defaultdict(list)
        bfc_alg  = defaultdict(list)
        ic_alg   = defaultdict(list)
        fc_alg   = defaultdict(list)
        clb_alg  = defaultdict(list)
        cub_alg  = defaultdict(list)
        rc_alg   = defaultdict(list)
        for s in scales:
            print('[{:>9.2f}] Scale: {:>.3f} starts'.format(time_mark(), s))
            c.objective.set_linear(target_var0, s*c0)

            try:
                c.solve()
            except CplexSolverError:
                print("Exception raised during solve")
                return

            obj_alg.append( c.solution.get_objective_value() )
            for y in instance.time_optimize:
                key = str(y)
                target_var   = 'V_Capacity(' + target_tech + '_' + key + ')'
                coefficient  = c.objective.get_linear(target_var)
                if y != target_year:
                    if not validate_coef(coefficient, instance, target_tech, y):
                        print('Error!')
                        sys.exit(0)
                capacity = c.solution.get_values(target_var)
                try:
                    # Out of some unknow reason, sometimes this function will 
                    # fail even though the model is totally feasible.
                    # Notes: This function fails when cross-over is not selected
                    # when barrier algorithm is selected
                    c_bound = c.solution.sensitivity.objective(target_var)
                    s_be    = c_bound[0] / coefficient # Break-even scale
                except:
                    c_bound = [None, None]
                    s_be    = None
                cost_i = s*value( instance.CostInvest[target_tech, y] )
                cost_f = s*value( instance.CostFixed[y, target_tech, y] )
                

                cap_alg[key].append(capacity)
                coef_alg[key].append(coefficient)
                ic_alg[key].append(cost_i)
                fc_alg[key].append(cost_f)
                if s_be:
                    bic_alg[key].append(s_be*cost_i)
                    bfc_alg[key].append(s_be*cost_f)
                else:
                    bic_alg[key].append(None)
                    bfc_alg[key].append(None)
                clb_alg[key].append( c_bound[0] )
                cub_alg[key].append( c_bound[1] )
                rc_alg[key].append( c.solution.get_reduced_costs(target_var) )

        obj[algorithm]  = obj_alg
        cap[algorithm]  = cap_alg
        coef[algorithm] = coef_alg
        bic[algorithm]  = bic_alg
        bfc[algorithm]  = bfc_alg
        ic[algorithm]   = ic_alg
        fc[algorithm]   = fc_alg
        clb[algorithm]  = clb_alg
        cub[algorithm]  = cub_alg
        rc[algorithm]   = rc_alg

        # Write to Excel spreadsheet
        print('[{:>9.2f}] Saving to Excel spreadsheet'.format( time_mark() ))
        row_title = [
            'scale',       'obj',       'cap', 'clb', 'coef', 
            'cub',   'bic (clb)', 'bfc (clb)',  'ic',   'fc',
            'rc'
        ]
        wb = Workbook()
        # for ws_title in cap_alg:
        for year in all_periods:
            ws_title = str(year)
            if ws_title not in cap_alg:
                continue
            ws = wb.create_sheet(ws_title)

            row = [
                scales, 
                obj_alg, 
                cap_alg[ws_title], 
                clb_alg[ws_title], 
                coef_alg[ws_title], 
                cub_alg[ws_title], 
                bic_alg[ws_title], 
                bfc_alg[ws_title], 
                ic_alg[ws_title], 
                fc_alg[ws_title],
                rc_alg[ws_title]
            ]

            # Note Python starts from 0, but row number starts from 1
            for j in range(0, len(row_title) ):
                cell = ws.cell(row = 1, column = j + 1)
                cell.value = row_title[j]
            for i in range(0, len(scales)):
                for j in range(0, len(row_title)):
                    cell = ws.cell(row = i + 2, column = j + 1)
                    cell.value = row[j][i]
        fname = '.'.join(
            [target_tech, str(target_year)]
            + [ i[:-4] for i in list_dat ] # Remove the .dat extension
            + [algorithm]
        ) # tech_name.year.dat_file_name.algorithm.xlsx
        wb.save(fname + '.xlsx')

def sen_range(tech, vintage, scales, dat):
    # Given a range of scaling factor for coefficient of a specific V_Capacity, 
    # returns objective value, reduced cost, capacity etc. for each scaling 
    # factor
    from openpyxl import Workbook
    target_year = vintage
    target_tech = tech
    algmap = {
        'primal simplex': 1,
        'dual simplex':   2,
        'barrier':        4,
        'default':        0,
    } # cplex definition

    t0 = time()
    time_mark = lambda: time() - t0

    model = return_Temoa_model()
    data = return_Temoa_data(model, dat)
    optimizer = SolverFactory('cplex')

    ic0         = data['CostInvest'][target_tech, target_year]
    fc0         = data['CostFixed'][target_year, target_tech, target_year]
    all_periods = data['time_future']

    obj  = dict()
    cap  = dict()
    lrc  = dict()
    coef = dict()
    urc  = dict()
    bic  = dict()
    bfc  = dict()
    ic   = dict() # Original IC
    fc   = dict() # Original FC

    for algorithm in ['barrier', 'dual simplex', 'primal simplex']:
        optimizer.options['lpmethod'] = algmap[algorithm]
        print('Algorithm: {}'.format( algorithm ))

        obj_alg  = list()
        cap_alg  = defaultdict(list)
        lrc_alg  = defaultdict(list)
        coef_alg = defaultdict(list)
        urc_alg  = defaultdict(list)
        bic_alg  = defaultdict(list)
        bfc_alg  = defaultdict(list)
        ic_alg   = defaultdict(list)
        fc_alg   = defaultdict(list)
        for s in scales:
            print('[{:>9.2f}] Scale: {:>.3f} starts'.format(time_mark(), s))
            data['CostInvest'][target_tech, target_year] = s*ic0
            for y in data['time_future']:
                if (y, target_tech, target_year) in data['CostFixed']:
                    data['CostFixed'][y, target_tech, target_year] = s*fc0
            instance = model.create_instance(data)
            instance.preprocess()
            results = optimizer.solve(instance, suffixes=['dual', 'urc', 'slack', 'lrc'])
            instance.solutions.load_from(results)

            obj_alg.append( value(instance.TotalCost) )
            for y in instance.time_optimize:
                key = str(y)
                c_vector = return_c_vector(instance, [])
                coefficient = c_vector[ ( 'V_Capacity', (target_tech, y) )]
                capacity = value(instance.V_Capacity[target_tech, y])
                lower_rc = value(
                    instance.lrc[ instance.V_Capacity[target_tech, y] ]
                )
                upper_rc = value(
                    instance.urc[ instance.V_Capacity[target_tech, y] ]
                )
                cost_i   = value( instance.CostInvest[target_tech, y] )
                cost_f   = value( instance.CostFixed[y, target_tech, y] )
                s_be = ( coefficient - lower_rc ) / coefficient # Break-even scale

                cap_alg[key].append(capacity)
                lrc_alg[key].append(lower_rc)
                coef_alg[key].append(coefficient)
                urc_alg[key].append(upper_rc)
                ic_alg[key].append(cost_i)
                fc_alg[key].append(cost_f)
                bic_alg[key].append(s_be*cost_i)
                bfc_alg[key].append(s_be*cost_f)

        obj[algorithm]  = obj_alg
        cap[algorithm]  = cap_alg
        lrc[algorithm]  = lrc_alg
        coef[algorithm] = coef_alg
        urc[algorithm]  = urc_alg
        bic[algorithm]  = bic_alg
        bfc[algorithm]  = bfc_alg
        ic[algorithm]   = ic_alg
        fc[algorithm]   = fc_alg

        # Write to Excel spreadsheet
        print('[{:>9.2f}] Saving to Excel spreadsheet'.format( time_mark() ))
        row_title = [
            'scale', 'obj', 'cap', 'lrc', 'coef', 
            'urc',   'bic', 'bfc', 'ic',  'fc'
        ]
        wb = Workbook()
        # for ws_title in cap_alg:
        for year in all_periods:
            ws_title = str(year)
            if ws_title not in cap_alg:
                continue
            ws = wb.create_sheet(ws_title)

            row = [
                scales, 
                obj_alg, 
                cap_alg[ws_title], 
                lrc_alg[ws_title], 
                coef_alg[ws_title], 
                urc_alg[ws_title], 
                bic_alg[ws_title], 
                bfc_alg[ws_title], 
                ic_alg[ws_title], 
                fc_alg[ws_title]
            ]

            # Note Python starts from 0, but row number starts from 1
            for j in range(0, len(row_title) ):
                c = ws.cell(row = 1, column = j + 1)
                c.value = row_title[j]
            for i in range(0, len(scales)):
                for j in range(0, len(row_title)):
                    c = ws.cell(row = i + 2, column = j + 1)
                    c.value = row[j][i]
        fname = '.'.join(
            [target_tech, str(target_year)]
            + [ i[:-4] for i in dat ] # Remove the .dat extension
            + [algorithm]
        ) # tech_name.year.dat_file_name.algorithm.xlsx
        wb.save(fname + '.xlsx')

def explore_Cost_marginal(dat):

    model = temoa_create_model()
    
    model.dual  = Suffix(direction=Suffix.IMPORT)
    model.rc    = Suffix(direction=Suffix.IMPORT)
    model.slack = Suffix(direction=Suffix.IMPORT)
    model.lrc   = Suffix(direction=Suffix.IMPORT)
    model.urc   = Suffix(direction=Suffix.IMPORT)

    data = DataPortal(model = model)
    for d in dat:
        data.load(filename=d)
    instance = model.create_instance(data)

    # Deactivate the DemandActivity constraint
    # instance.DemandActivityConstraint.deactivate()
    # instance.preprocess()

    optimizer = SolverFactory('cplex')
    results = optimizer.solve(
        instance, 
        keepfiles=True,
        suffixes=['dual', 'urc', 'slack', 'lrc']
        )
    instance.solutions.load_from(results)

    print('Dual and slack variables for emission caps:')
    for e in instance.commodity_emissions:
        for p in instance.time_optimize:
            if (p, e) in instance.EmissionLimitConstraint:
                print(p, e, instance.dual[instance.EmissionLimitConstraint[p, e]], '\t', instance.slack[instance.EmissionLimitConstraint[p, e]])

    # print('Dual and slack variables for Commodity Demand Constraints')
    # for c in instance.commodity_demand:
    #     for p in instance.time_optimize:
    #         for s in instance.time_season:
    #             for tod in instance.time_of_day:
    #                 print(p, s, tod, instance.dual[instance.DemandConstraint[p,s,tod,c]], instance.slack[instance.DemandConstraint[p,s,tod,c]])

def plot_breakeven(years, bic, ic):
    # bic is a dictionary, ic is a list of the raw investment costs
    # ic = [x, x, ..., x], the length of which equals to the length of years
    # bic[scenario] = [x, x, x... x] where the length equals to the number 
    # of optimized periods.
    sen_color_map = {
        'IC':    [0.9, 0.9, 0.9],
        'LF':    'black',
        'R':     'black',
        'HF':    'black',
        'HD':    'black',
        'CPPLF': 'green',
        'CPP':   'green',
        'CPPHF': 'green',
        'CPPHD': 'green'
    }

    sen_lstyle_map = {
        'IC':    None,
        'LF':    '--',
        'R':     '-',
        'HF':    '-.',
        'HD':    ':',
        'CPPLF': '--',
        'CPP':   '-',
        'CPPHF': '-.',
        'CPPHD': ':'
    }

    sen_marker_map = {
        'IC':    None,
        'LF':    's',
        'R':     's',
        'HF':    's',
        'HD':    's',
        'CPPLF': 's',
        'CPP':   's',
        'CPPHF': 's',
        'CPPHD': 's'
    }

    scenarios = bic.keys()
    h = plt.figure()
    ax = plt.subplot(111)
    ax.fill_between(years, 0, ic, 
        facecolor = sen_color_map['IC']
        )

    for s in scenarios:
        ax.plot(years, bic[s], 
            color = sen_color_map[s],
            # marker = sen_marker_map[s],
            linestyle = sen_lstyle_map[s]
            )
    ax.yaxis.grid(True)
    plt.ylabel('$/MWh')
    plt.xlim( ( years[0]-5, years[-1]+5 ) )
    return ax

def bin_search_and_range():
    def return_range(bs):
        # Given break-even scaling factor, return an appropriate range
        ub = None
        lb = None
        if bs >= 0.85:
            lb = int(bs*100) - 15
            ub = 100
        elif bs <= 0.15:
            lb = 1
            ub = int(bs*100) + 15
        else:
            lb = int(bs*100) - 15
            ub = int(bs*100) + 15
        return [0.001*i for i in range(lb*10, ub*10, 10)]

    list_file = ['reference.dat', 'NCupdated_noLeadTime.dat']
    list_tech = ['EBIOIGCC', 'EURNALWR15']
    monitor_vintage = 2020
    eps = 0.01
    for f in list_file:
        for t in list_tech:
            bs = bin_search(t, monitor_vintage, [f], eps)
            sen_range( t, monitor_vintage, return_range(bs), [f] )

if __name__ == "__main__":
    # sen_bin_search(
    #     'ECOALIGCCS', 
    #     2020,
    #     ['reference.dat'],
    #     0.01
    # )
    scales = [0.001* i for i in range(250, 260, 10)]
    sen_range(
        'ECOALIGCCS', 
        2020,
        scales,
        ['reference.dat']
    )
    # do_sensitivity_new()
    # do_sensitivity_old()
    # explore_Cost_marginal(['/afs/unity.ncsu.edu/users/b/bli6/TEMOA_NC/sql20170417/results/R/NCreference.R.dat'])
